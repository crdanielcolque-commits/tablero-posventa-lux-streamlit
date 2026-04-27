# ============================================================
# TABLERO POSVENTA — MACRO → MICRO (Semanal + Acumulado)
# v2.3.26 — Dirección + Gestión Inteligente
# + Filtros obligatorios con MULTISELECCIÓN:
#   - Mes
#   - Semana corte
#   - Sucursal
# + Chips de multiselect en VERDE (no rojo)
# + Fix StreamlitDuplicateElementId en sparklines P&L
# + Export Excel profesional respetando filtros múltiples
# + 3 tabs operativos:
#   🔧 Órdenes Abiertas
#   🧾 Pend. Facturación
#   💬 Presupuestos
# ============================================================

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import gdown
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------
# CONFIG
# ---------------------------
st.set_page_config(page_title="Tablero Posventa", layout="wide")

DRIVE_FILE_ID = "191JKfQWj3yehcnisKTPDs_KpWaOTyslhQ0g273Xvzjc"
EXCEL_LOCAL = "base_posventa.xlsx"

SUCURSAL_MAP = {
    2: "Jujuy",
    3: "Taller Móvil",
    4: "Salta",
    5: "Tartagal",
    7: "Lajitas",
    9: "Chapa y Pintura",
}

# ---------------------------
# HELPERS GENERALES
# ---------------------------
def parse_semana_num(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    num = s.str.extract(r"(\d+(?:[.,]\d+)?)")[0]
    num = num.str.replace(",", ".", regex=False)
    numf = pd.to_numeric(num, errors="coerce")
    return np.floor(numf).astype("Int64")

def to_num_ar(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return np.nan
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return np.nan

    is_pct = "%" in s
    s = s.replace("%", "")
    s = (
        s.replace("$", "")
         .replace("AR$", "")
         .replace(" ", "")
         .replace("\u00A0", "")
    )

    if "," in s and "." in s:
        last_comma = s.rfind(",")
        last_dot = s.rfind(".")
        if last_dot > last_comma:
            s = s.replace(",", "")
        else:
            s = s.replace(".", "")
            s = s.replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        pass

    try:
        v = float(s)
        if is_pct:
            v = v / 100.0
        return v
    except Exception:
        return np.nan

def safe_ratio(n, d):
    try:
        if d is None or pd.isna(d) or float(d) == 0:
            return np.nan
        return float(n) / float(d)
    except Exception:
        return np.nan

def money_str(x):
    if x is None or pd.isna(x):
        return "—"
    try:
        return f"${float(x):,.0f}".replace(",", ".")
    except Exception:
        return "—"

def qty_str(x):
    if x is None or pd.isna(x):
        return "—"
    try:
        return f"{float(x):,.0f}".replace(",", ".")
    except Exception:
        return "—"

def pct_str(x):
    if x is None or pd.isna(x):
        return "—"
    try:
        return f"{float(x)*100:.1f}%"
    except Exception:
        return "—"

def card_html_base(title, value, sub):
    return f"""
    <div style="border:1px solid #eee;border-radius:14px;padding:16px;background:#fff;
                box-shadow:0 2px 10px rgba(0,0,0,0.04);">
        <div style="font-size:12px;color:#6c757d;font-weight:800;letter-spacing:0.2px;">{title}</div>
        <div style="font-size:28px;font-weight:900;margin-top:6px;">{value}</div>
        <div style="font-size:12px;color:#6c757d;margin-top:6px;">{sub}</div>
    </div>
    """

def hide_sidebar_css():
    return """
    <style>
      section[data-testid="stSidebar"] {display: none !important;}
      div[data-testid="stSidebarNav"] {display: none !important;}
      .block-container {padding-left: 2.2rem; padding-right: 2.2rem;}
    </style>
    """

def multiselect_green_tags_css():
    return """
    <style>
    div[data-baseweb="select"] span[data-baseweb="tag"]{
        background-color: #16a34a !important;
        border-color: #16a34a !important;
    }
    div[data-baseweb="select"] span[data-baseweb="tag"] *,
    div[data-baseweb="select"] span[data-baseweb="tag"] span{
        color: #ffffff !important;
    }
    div[data-baseweb="select"] span[data-baseweb="tag"] svg{
        fill: #ffffff !important;
        color: #ffffff !important;
    }
    div[data-baseweb="select"] span[data-baseweb="tag"] svg:hover{
        fill: #eaffea !important;
        color: #eaffea !important;
    }
    </style>
    """

st.markdown(multiselect_green_tags_css(), unsafe_allow_html=True)

def norm_text(x: str) -> str:
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = (s.replace("á","a").replace("é","e").replace("í","i")
           .replace("ó","o").replace("ú","u").replace("ü","u")
           .replace("ñ","n"))
    s = s.replace("_", " ").replace("-", " ")
    s = " ".join(s.split())
    return s

def month_name_es(month_num: int) -> str:
    m = {
        1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio",
        7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"
    }
    return m.get(int(month_num), "")

def compute_semana_mes(df_in: pd.DataFrame) -> pd.Series:
    out = pd.Series(index=df_in.index, dtype="Int64")
    for mes, g in df_in.groupby("Mes"):
        uniq = sorted([int(x) for x in g["Semana_Num"].dropna().unique().tolist()])
        mapping = {w: i+1 for i, w in enumerate(uniq)}
        out.loc[g.index] = g["Semana_Num"].map(mapping).astype("Int64")
    return out

def find_sheet_name(sheet_names, target: str):
    t = norm_text(target)
    for s in sheet_names:
        if norm_text(s) == t:
            return s
    for s in sheet_names:
        if t in norm_text(s):
            return s
    t_words = set(norm_text(target).split())
    best = None
    best_score = 0
    for s in sheet_names:
        sw = set(norm_text(s).split())
        score = len(t_words.intersection(sw))
        if score > best_score:
            best_score = score
            best = s
    return best if best_score >= max(2, len(t_words)//2) else None

def build_month_label(mes_key: str) -> str:
    try:
        p = pd.Period(mes_key, freq="M")
        return f"{month_name_es(p.month)} {p.year}"
    except Exception:
        return str(mes_key)

def labels_from_mes_keys(mes_keys: list[str]) -> str:
    if not mes_keys:
        return "—"
    labels = [build_month_label(m) for m in mes_keys]
    if len(labels) == 1:
        return labels[0]
    return " | ".join(labels)

def list_to_export_text(values):
    if not values:
        return "TODOS"
    return " | ".join([str(v) for v in values])

# ---------------------------
# HELPERS EXCEL PROFESIONAL
# ---------------------------
def _autosize_ws(ws, min_w=10, max_w=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = max(min_w, min(max_w, max_len + 2))
        ws.column_dimensions[col_letter].width = width

def _style_table(ws, freeze_row=1):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.freeze_panes = ws[f"A{freeze_row+1}"]
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = left

def _format_columns(ws, col_formats: dict):
    headers = [c.value for c in ws[1]]
    name_to_idx = {str(h): i+1 for i, h in enumerate(headers) if h is not None}
    for col_name, fmt in col_formats.items():
        if col_name in name_to_idx:
            idx = name_to_idx[col_name]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = fmt

def build_exec_excel_professional(
    meta_df: pd.DataFrame,
    resumen_df: pd.DataFrame,
    pl_sucursal_rep: pd.DataFrame,
    pl_sucursal_srv: pd.DataFrame,
    pl_ap_rep: pd.DataFrame,
    pl_ap_srv: pd.DataFrame,
    ranking_rep: pd.DataFrame,
    ranking_srv: pd.DataFrame,
    hitos_df: pd.DataFrame,
    abiertas_df: pd.DataFrame,
    pendientes_fact_df: pd.DataFrame,
    presup_df: pd.DataFrame,
) -> BytesIO:
    def clean(df_in: pd.DataFrame, keep_cols=None, drop_cols=None):
        if df_in is None:
            return pd.DataFrame()
        dfc = df_in.copy()
        if drop_cols:
            for c in drop_cols:
                if c in dfc.columns:
                    dfc = dfc.drop(columns=[c])
        if keep_cols:
            cols = [c for c in keep_cols if c in dfc.columns]
            dfc = dfc[cols].copy()
        return dfc

    resumen_clean = clean(
        resumen_df,
        keep_cols=["Bloque","Real_Acum","Obj_Acum","Cumpl_Acum","Proy_EOM_RunRate","Dias_Transc","Dias_Mes"]
    )

    rep_s = clean(pl_sucursal_rep, drop_cols=["label","Cumpl_plot"])
    srv_s = clean(pl_sucursal_srv, drop_cols=["label","Cumpl_plot"])
    if not rep_s.empty:
        rep_s.insert(0, "Bloque", "Repuestos")
    if not srv_s.empty:
        srv_s.insert(0, "Bloque", "Servicios")
    pl_suc = pd.concat([rep_s, srv_s], ignore_index=True) if (not rep_s.empty or not srv_s.empty) else pd.DataFrame()

    rep_a = clean(pl_ap_rep, drop_cols=["label","Cumpl_plot"])
    srv_a = clean(pl_ap_srv, drop_cols=["label","Cumpl_plot"])
    if not rep_a.empty:
        rep_a.insert(0, "Bloque", "Repuestos")
    if not srv_a.empty:
        srv_a.insert(0, "Bloque", "Servicios")
    pl_ap = pd.concat([rep_a, srv_a], ignore_index=True) if (not rep_a.empty or not srv_a.empty) else pd.DataFrame()

    rep_r = clean(ranking_rep, drop_cols=["label","Cumpl_plot","key"])
    srv_r = clean(ranking_srv, drop_cols=["label","Cumpl_plot","key"])
    if not rep_r.empty:
        rep_r.insert(0, "Bloque", "Repuestos")
    if not srv_r.empty:
        srv_r.insert(0, "Bloque", "Servicios")
    rank = pd.concat([rep_r, srv_r], ignore_index=True) if (not rep_r.empty or not srv_r.empty) else pd.DataFrame()

    abiertas_xls = clean(abiertas_df)
    pf_xls = clean(pendientes_fact_df)
    presup_xls = clean(presup_df)
    hitos = pd.DataFrame() if hitos_df is None else hitos_df.copy()

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        meta_df.to_excel(writer, index=False, sheet_name="00_Meta")
        resumen_clean.to_excel(writer, index=False, sheet_name="01_Resumen")
        pl_suc.to_excel(writer, index=False, sheet_name="02_P&L_Sucursal")
        pl_ap.to_excel(writer, index=False, sheet_name="03_P&L_Aperturas")
        rank.to_excel(writer, index=False, sheet_name="04_Ranking_Micro")
        hitos.to_excel(writer, index=False, sheet_name="05_Hitos_mes")
        abiertas_xls.to_excel(writer, index=False, sheet_name="06_Abiertas")
        pf_xls.to_excel(writer, index=False, sheet_name="07_Pend_Fact")
        presup_xls.to_excel(writer, index=False, sheet_name="08_Presup")

    out.seek(0)
    wb = load_workbook(out)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        _style_table(ws)
        _autosize_ws(ws)

    _format_columns(wb["01_Resumen"], {
        "Real_Acum": '"$"#,##0',
        "Obj_Acum": '"$"#,##0',
        "Cumpl_Acum": '0.0%',
        "Proy_EOM_RunRate": '"$"#,##0',
        "Dias_Transc": '0',
        "Dias_Mes": '0',
    })

    for ws_name in ["02_P&L_Sucursal", "03_P&L_Aperturas", "04_Ranking_Micro"]:
        _format_columns(wb[ws_name], {
            "Real": '"$"#,##0',
            "Obj": '"$"#,##0',
            "Cumpl": '0.0%',
        })

    for ws_name in ["06_Abiertas", "07_Pend_Fact", "08_Presup"]:
        if ws_name in wb.sheetnames:
            _format_columns(wb[ws_name], {
                "Imp_Cliente": '"$"#,##0',
                "Imp_Interna": '"$"#,##0',
                "Imp_Garantia": '"$"#,##0',
                "Monto": '"$"#,##0',
                "Antig_Dias": '0',
            })

    out2 = BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2

# ---------------------------
# HELPERS OPERATIVOS
# ---------------------------
def detect_first_matching_column(df: pd.DataFrame, keywords: list[str]):
    cols = list(df.columns)
    norm_cols = {c: norm_text(c) for c in cols}
    for kw in keywords:
        nkw = norm_text(kw)
        for c, nc in norm_cols.items():
            if nkw == nc or nkw in nc:
                return c
    return None

def map_sucursal_codes(series: pd.Series) -> pd.Series:
    nums = pd.to_numeric(series, errors="coerce")
    return nums.map(SUCURSAL_MAP).fillna(series.astype(str))

def build_operational_standard(df_raw: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    cols_out = [
        "Sucursal","Nro de Orden","Cliente","Patente","Fecha",
        "Antig_Dias","Imp_Cliente","Imp_Interna","Imp_Garantia","Monto",
        "Asesor","Estado","Origen"
    ]

    if df_raw is None or df_raw.empty:
        return pd.DataFrame(columns=cols_out)

    df = df_raw.copy()
    df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")]
    if df.empty:
        return pd.DataFrame(columns=cols_out)

    sheet_norm = norm_text(sheet_name)

    suc_col = None
    if sheet_norm == "presupuestos" and len(df.columns) >= 5:
        suc_col = df.columns[4]
    if suc_col is None:
        suc_col = detect_first_matching_column(df, ["Suc.", "Sucursal", "Suc", "SUCURSAL"])
    if suc_col is None and len(df.columns) >= 1:
        suc_col = df.columns[0]

    doc_col = None
    if sheet_norm in ["abiertas", "pendientes fact", "pendientes facturacion", "pend fact"] and len(df.columns) >= 5:
        doc_col = df.columns[4]
    if doc_col is None:
        doc_col = detect_first_matching_column(df, [
            "Ord.Rep.", "Ord Rep", "OT", "Orden", "Nro OT", "Numero OT", "N° OT",
            "Nro", "Numero", "Presupuesto", "Nro Presupuesto"
        ])

    cliente_col = detect_first_matching_column(df, [
        "Cliente", "Apellido y Nombre", "Nombre", "Razon Social"
    ])

    patente_col = detect_first_matching_column(df, ["Patente", "Dominio"])

    fecha_col = None
    if sheet_norm == "abiertas" and len(df.columns) >= 2:
        fecha_col = df.columns[1]
    elif sheet_norm in ["pendientes fact", "pendientes facturacion", "pend fact"] and len(df.columns) >= 3:
        fecha_col = df.columns[2]
    if fecha_col is None:
        fecha_col = detect_first_matching_column(df, [
            "Apertura", "Fecha", "Fecha Apertura", "Fecha Ingreso", "Ingreso",
            "Fecha Emision", "Fecha Presupuesto", "Fecha OT", "Alta", "Emision"
        ])

    asesor_col = None
    if sheet_norm in ["abiertas", "pendientes fact", "pendientes facturacion", "pend fact"] and len(df.columns) >= 15:
        asesor_col = df.columns[14]
    if asesor_col is None:
        asesor_col = detect_first_matching_column(df, [
            "Asesor", "Recepcionista", "Asesor Servicio", "Responsable", "Vendedor"
        ])

    estado_col = None
    if sheet_norm == "presupuestos" and len(df.columns) >= 8:
        estado_col = df.columns[7]
    if estado_col is None:
        estado_col = detect_first_matching_column(df, ["Estado", "Situacion", "Situación", "Status"])

    out = pd.DataFrame()
    out["Sucursal"] = map_sucursal_codes(df[suc_col]) if suc_col in df.columns else "—"
    out["Nro de Orden"] = df[doc_col].astype(str) if doc_col in df.columns else ""
    out["Cliente"] = df[cliente_col].astype(str) if cliente_col in df.columns else ""
    out["Patente"] = df[patente_col].astype(str) if patente_col in df.columns else ""
    out["Fecha"] = pd.to_datetime(df[fecha_col], errors="coerce") if fecha_col in df.columns else pd.NaT
    out["Asesor"] = df[asesor_col].astype(str) if asesor_col in df.columns else ""
    out["Estado"] = df[estado_col].astype(str) if estado_col in df.columns else ""
    out["Origen"] = sheet_name

    out["Imp_Cliente"] = np.nan
    out["Imp_Interna"] = np.nan
    out["Imp_Garantia"] = np.nan
    out["Monto"] = np.nan

    if sheet_norm in ["abiertas", "pendientes fact", "pendientes facturacion", "pend fact"]:
        if len(df.columns) >= 14:
            col_l = df.columns[11]
            col_m = df.columns[12]
            col_n = df.columns[13]

            out["Imp_Cliente"] = df[col_l].apply(to_num_ar)
            out["Imp_Interna"] = df[col_m].apply(to_num_ar)
            out["Imp_Garantia"] = df[col_n].apply(to_num_ar)
            out["Monto"] = (
                out["Imp_Cliente"].fillna(0)
                + out["Imp_Interna"].fillna(0)
                + out["Imp_Garantia"].fillna(0)
            )
            out["Monto"] = out["Monto"].where(out["Monto"] != 0, np.nan)
    else:
        monto_col = detect_first_matching_column(df, [
            "Monto", "Importe", "Total", "Presupuesto",
            "Total Presupuesto", "Saldo", "Importe Total"
        ])
        if monto_col in df.columns:
            out["Monto"] = df[monto_col].apply(to_num_ar)

    today = pd.Timestamp.today().normalize()
    out["Antig_Dias"] = (today - out["Fecha"]).dt.days
    out["Antig_Dias"] = out["Antig_Dias"].where(out["Antig_Dias"] >= 0)

    for c in ["Nro de Orden", "Cliente", "Patente", "Asesor", "Estado"]:
        out[c] = out[c].replace("nan", "").fillna("").astype(str).str.strip()

    keep_mask = (
        out["Nro de Orden"].ne("") |
        out["Cliente"].ne("") |
        out["Patente"].ne("") |
        out["Fecha"].notna() |
        out["Monto"].notna() |
        out["Imp_Cliente"].notna() |
        out["Imp_Interna"].notna() |
        out["Imp_Garantia"].notna()
    )
    out = out[keep_mask].copy()

    return out

def add_age_bucket(df_std: pd.DataFrame) -> pd.DataFrame:
    x = df_std.copy()
    if "Antig_Dias" not in x.columns:
        x["Age_Bucket"] = "Sin fecha"
        return x

    def bucket(v):
        if pd.isna(v):
            return "Sin fecha"
        v = float(v)
        if v <= 2:
            return "0-2 días"
        if v <= 5:
            return "3-5 días"
        if v <= 10:
            return "6-10 días"
        if v <= 15:
            return "11-15 días"
        return "16+ días"

    x["Age_Bucket"] = x["Antig_Dias"].apply(bucket)
    return x

def op_summary(df_std: pd.DataFrame):
    if df_std is None or df_std.empty:
        return {"count": 0, "monto": np.nan, "age_avg": np.nan, "age_max": np.nan}
    return {
        "count": int(len(df_std)),
        "monto": df_std["Monto"].sum(min_count=1) if "Monto" in df_std.columns else np.nan,
        "age_avg": df_std["Antig_Dias"].mean() if "Antig_Dias" in df_std.columns else np.nan,
        "age_max": df_std["Antig_Dias"].max() if "Antig_Dias" in df_std.columns else np.nan,
    }

def format_money_cols(df_in: pd.DataFrame, cols_money: list[str]) -> pd.DataFrame:
    df = df_in.copy()
    for c in cols_money:
        if c in df.columns:
            df[c] = df[c].apply(money_str)
    return df

def render_operational_tab(
    df_std: pd.DataFrame,
    title: str,
    purpose_text: str,
    key_prefix: str,
    enable_asesor_filter: bool = False
):
    st.markdown(f"## {title}")
    st.caption(purpose_text)
    st.markdown("---")

    if df_std is None or df_std.empty:
        st.info("No hay registros en esta hoja o no se pudieron interpretar columnas con contenido útil.")
        return

    base = add_age_bucket(df_std.copy())

    sucursales_disp = sorted([s for s in base["Sucursal"].dropna().astype(str).str.strip().unique().tolist() if s != ""])
    colf1, colf2 = st.columns([1.2, 1.8])
    with colf1:
        suc_sel = st.multiselect(
            "Filtrar sucursal",
            sucursales_disp,
            default=sucursales_disp,
            key=f"{key_prefix}_suc"
        ) if sucursales_disp else []
    with colf2:
        st.caption("Todos los gráficos y tablas respetan este filtro.")

    x0 = base[base["Sucursal"].astype(str).isin(suc_sel)].copy() if sucursales_disp else base.copy()

    if enable_asesor_filter:
        asesores_disp = sorted([a for a in x0["Asesor"].dropna().astype(str).str.strip().unique().tolist() if a != ""])
        f1, f2 = st.columns([1.6, 1.4])

        with f1:
            asesor_sel = st.multiselect(
                "Filtrar recepcionista / asesor",
                asesores_disp,
                default=asesores_disp,
                key=f"{key_prefix}_asesor"
            ) if asesores_disp else []

        with f2:
            st.caption("Los gráficos y tablas de abajo respetan este filtro.")

        if asesores_disp:
            x = x0[x0["Asesor"].astype(str).isin(asesor_sel)].copy()
        else:
            x = x0.copy()

        st.markdown("### Participación por recepcionista")
        p1, p2 = st.columns(2)

        ga_count = (
            x0.groupby("Asesor", as_index=False)
            .agg(Casos=("Asesor", "count"))
            .sort_values("Casos", ascending=False)
        )
        ga_count = ga_count[ga_count["Asesor"].astype(str).str.strip() != ""]

        with p1:
            st.markdown("**% sobre cantidad de órdenes**")
            if ga_count.empty:
                st.info("Sin recepcionistas identificados.")
            else:
                fig = px.pie(ga_count, names="Asesor", values="Casos")
                fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
                st.plotly_chart(fig, use_container_width=True, key=f"{key_prefix}_pie_cantidad")

        ga_monto = (
            x0.groupby("Asesor", as_index=False)
            .agg(Monto=("Monto", "sum"))
            .sort_values("Monto", ascending=False)
        )
        ga_monto = ga_monto[
            (ga_monto["Asesor"].astype(str).str.strip() != "") &
            (ga_monto["Monto"].fillna(0) > 0)
        ]

        with p2:
            st.markdown("**% sobre valor económico**")
            if ga_monto.empty:
                st.info("Sin importes para distribuir por recepcionista.")
            else:
                fig = px.pie(ga_monto, names="Asesor", values="Monto")
                fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
                st.plotly_chart(fig, use_container_width=True, key=f"{key_prefix}_pie_monto")
    else:
        x = x0.copy()

    summ = op_summary(x)

    imp_cliente = x["Imp_Cliente"].sum(min_count=1) if "Imp_Cliente" in x.columns else np.nan
    imp_interna = x["Imp_Interna"].sum(min_count=1) if "Imp_Interna" in x.columns else np.nan
    imp_garantia = x["Imp_Garantia"].sum(min_count=1) if "Imp_Garantia" in x.columns else np.nan
    monto_total = x["Monto"].sum(min_count=1) if "Monto" in x.columns else np.nan

    mix_cliente = safe_ratio(imp_cliente, monto_total)
    mix_interna = safe_ratio(imp_interna, monto_total)
    mix_garantia = safe_ratio(imp_garantia, monto_total)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(card_html_base("Cantidad", qty_str(summ["count"]), "Casos totales"), unsafe_allow_html=True)
    with c2:
        st.markdown(card_html_base("Monto potencial", money_str(summ["monto"]), "Monto total"), unsafe_allow_html=True)
    with c3:
        st.markdown(card_html_base("Antig. prom.", qty_str(round(summ["age_avg"], 0) if pd.notna(summ["age_avg"]) else np.nan), "Días promedio"), unsafe_allow_html=True)
    with c4:
        st.markdown(card_html_base("Antig. máxima", qty_str(summ["age_max"]), "Días"), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Mix de facturación")

    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(card_html_base("Mix Cliente", pct_str(mix_cliente), f"Monto {money_str(imp_cliente)}"), unsafe_allow_html=True)
    with m2:
        st.markdown(card_html_base("Mix Interna", pct_str(mix_interna), f"Monto {money_str(imp_interna)}"), unsafe_allow_html=True)
    with m3:
        st.markdown(card_html_base("Mix Garantía", pct_str(mix_garantia), f"Monto {money_str(imp_garantia)}"), unsafe_allow_html=True)

    st.markdown("---")
    a, b = st.columns(2)

    with a:
        st.markdown("### Casos por sucursal")
        g = x.groupby("Sucursal", as_index=False).agg(
            Casos=("Sucursal", "count"),
            Monto=("Monto", "sum"),
            Antig_Prom=("Antig_Dias", "mean"),
            Imp_Cliente=("Imp_Cliente", "sum"),
            Imp_Interna=("Imp_Interna", "sum"),
            Imp_Garantia=("Imp_Garantia", "sum"),
        ).sort_values("Casos", ascending=False)

        fig = px.bar(g, x="Casos", y="Sucursal", orientation="h", text="Casos")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
        fig.update_traces(textposition="inside")
        st.plotly_chart(fig, use_container_width=True, key=f"{key_prefix}_bar_sucursal")

    with b:
        st.markdown("### Antigüedad")
        ga = x.groupby("Age_Bucket", as_index=False).agg(Casos=("Age_Bucket", "count"))
        order = ["0-2 días", "3-5 días", "6-10 días", "11-15 días", "16+ días", "Sin fecha"]
        ga["Age_Bucket"] = pd.Categorical(ga["Age_Bucket"], categories=order, ordered=True)
        ga = ga.sort_values("Age_Bucket")

        fig = px.bar(ga, x="Age_Bucket", y="Casos", text="Casos")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="")
        fig.update_traces(textposition="outside")
        st.plotly_chart(fig, use_container_width=True, key=f"{key_prefix}_bar_antig")

    st.markdown("---")
    t1, t2 = st.columns(2)

    cols = [c for c in [
        "Sucursal", "Asesor", "Nro de Orden", "Cliente", "Patente", "Fecha", "Antig_Dias",
        "Imp_Cliente", "Imp_Interna", "Imp_Garantia", "Monto", "Estado"
    ] if c in x.columns]

    with t1:
        st.markdown("### Top más antiguos")
        oldest = x.sort_values(["Antig_Dias", "Monto"], ascending=[False, False]).head(15)[cols].copy()
        oldest = format_money_cols(oldest, ["Imp_Cliente", "Imp_Interna", "Imp_Garantia", "Monto"])
        st.dataframe(oldest, use_container_width=True, hide_index=True)

    with t2:
        if "Monto" in x.columns and x["Monto"].notna().any():
            st.markdown("### Top mayor importe")
            biggest = x.sort_values(["Monto", "Antig_Dias"], ascending=[False, False]).head(15)[cols].copy()
            biggest = format_money_cols(biggest, ["Imp_Cliente", "Imp_Interna", "Imp_Garantia", "Monto"])
            st.dataframe(biggest, use_container_width=True, hide_index=True)
        else:
            st.markdown("### Vista adicional")
            latest = x.sort_values(["Fecha", "Antig_Dias"], ascending=[False, False]).head(15)[cols].copy()
            latest = format_money_cols(latest, ["Imp_Cliente", "Imp_Interna", "Imp_Garantia", "Monto"])
            st.dataframe(latest, use_container_width=True, hide_index=True)

    st.markdown("---")
    with st.expander("🔎 Detalle completo", expanded=False):
        cols = [c for c in [
            "Sucursal", "Asesor", "Nro de Orden", "Cliente", "Patente", "Fecha", "Antig_Dias",
            "Imp_Cliente", "Imp_Interna", "Imp_Garantia", "Monto", "Estado", "Origen"
        ] if c in x.columns]
        det = x[cols].copy()
        det = format_money_cols(det, ["Imp_Cliente", "Imp_Interna", "Imp_Garantia", "Monto"])
        st.dataframe(det, use_container_width=True, hide_index=True)

# ---------------------------
# LOAD DRIVE
# ---------------------------
@st.cache_data(ttl=300)
def load_from_drive():
    url = f"https://docs.google.com/spreadsheets/d/{DRIVE_FILE_ID}/export?format=xlsx"
    gdown.download(url, EXCEL_LOCAL, quiet=True)

    xls = pd.ExcelFile(EXCEL_LOCAL)
    sheet_names = list(xls.sheet_names)

    df0 = pd.read_excel(xls, sheet_name=0)
    df0 = df0.loc[:, ~df0.columns.astype(str).str.match(r"^Unnamed")]

    dias_sheet = find_sheet_name(sheet_names, "Dias habiles")
    if dias_sheet is not None:
        df_dias = pd.read_excel(xls, sheet_name=dias_sheet)
        df_dias = df_dias.loc[:, ~df_dias.columns.astype(str).str.match(r"^Unnamed")]
    else:
        df_dias = pd.DataFrame(columns=["Mes", "Semana", "Dias habiles"])

    resumen_sheet = find_sheet_name(sheet_names, "Resumen del mes")
    if resumen_sheet is not None:
        df_res = pd.read_excel(xls, sheet_name=resumen_sheet)
        df_res = df_res.loc[:, ~df_res.columns.astype(str).str.match(r"^Unnamed")]
    else:
        df_res = pd.DataFrame()

    abiertas_sheet = find_sheet_name(sheet_names, "Abiertas")
    pendientes_fact_sheet = find_sheet_name(sheet_names, "Pendientes Fact")
    presup_sheet = find_sheet_name(sheet_names, "Presupuestos")

    df_abiertas = pd.read_excel(xls, sheet_name=abiertas_sheet) if abiertas_sheet else pd.DataFrame()
    df_pfact = pd.read_excel(xls, sheet_name=pendientes_fact_sheet) if pendientes_fact_sheet else pd.DataFrame()
    df_presup = pd.read_excel(xls, sheet_name=presup_sheet) if presup_sheet else pd.DataFrame()

    return (
        df0, df_dias, df_res, sheet_names, dias_sheet, resumen_sheet,
        df_abiertas, df_pfact, df_presup,
        abiertas_sheet, pendientes_fact_sheet, presup_sheet
    )

(
    df, df_dias_habiles, df_resumen_mes, SHEET_NAMES, DIAS_SHEET_FOUND, RESUMEN_SHEET_FOUND,
    df_abiertas_raw, df_pfact_raw, df_presup_raw,
    ABIERTAS_SHEET_FOUND, PFACT_SHEET_FOUND, PRESUP_SHEET_FOUND
) = load_from_drive()

# ---------------------------
# VALIDACIÓN BASE
# ---------------------------
required = [
    "Fecha", "Semana", "Sucursal", "KPI", "Categoria_KPI", "Tipo_KPI",
    "Real_$", "Real_Q", "Objetivo_$", "Objetivo_Q", "Cumplimiento_%", "Estado"
]
missing = [c for c in required if c not in df.columns]
if missing:
    st.error("❌ Faltan columnas requeridas en el Excel:")
    st.write(missing)
    st.stop()

# ---------------------------
# NORMALIZACIÓN BASE
# ---------------------------
df["Semana_Num"] = parse_semana_num(df["Semana"])
df = df[~df["Semana_Num"].isna()].copy()

df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
df = df[~df["Fecha"].isna()].copy()

df["Mes"] = df["Fecha"].dt.to_period("M").astype(str)
df["Mes_Nombre"] = df["Fecha"].dt.month.apply(month_name_es)
df["Mes_norm"] = df["Mes_Nombre"].apply(norm_text)
df["Semana_Mes"] = compute_semana_mes(df)

for c in ["Real_$", "Costo_$", "Margen_$", "Margen_%", "Real_Q", "Objetivo_$", "Objetivo_Q", "Cumplimiento_%"]:
    if c in df.columns:
        df[c] = df[c].apply(to_num_ar)

df["KPI"] = df["KPI"].astype(str).str.strip()
df["Categoria_KPI"] = df["Categoria_KPI"].astype(str).str.strip()
df["Tipo_KPI"] = df["Tipo_KPI"].astype(str).str.strip()
df["Sucursal"] = df["Sucursal"].astype(str).str.strip()

def build_real_obj(row):
    if row["Tipo_KPI"] == "$":
        return row["Real_$"], row["Objetivo_$"]
    return row["Real_Q"], row["Objetivo_Q"]

tmp = df.apply(build_real_obj, axis=1, result_type="expand")
df["Real_val"] = pd.to_numeric(tmp[0], errors="coerce").fillna(0.0)
df["Obj_val"] = pd.to_numeric(tmp[1], errors="coerce").fillna(0.0)

if df_dias_habiles is None or df_dias_habiles.empty:
    df_dias_habiles = pd.DataFrame(columns=["Mes", "Semana", "Dias habiles"])
for col in ["Mes", "Semana", "Dias habiles"]:
    if col not in df_dias_habiles.columns:
        df_dias_habiles[col] = np.nan

df_dias_habiles["Mes_norm"] = df_dias_habiles["Mes"].apply(norm_text)
df_dias_habiles["Semana"] = pd.to_numeric(df_dias_habiles["Semana"], errors="coerce").fillna(0).astype(int)
df_dias_habiles["Dias habiles"] = pd.to_numeric(df_dias_habiles["Dias habiles"], errors="coerce").fillna(0).astype(float)

# ---------------------------
# CONTROLES VISUALES
# ---------------------------
if "modo_presentacion" not in st.session_state:
    st.session_state["modo_presentacion"] = False
if "cap_visual" not in st.session_state:
    st.session_state["cap_visual"] = True
if "cap_val" not in st.session_state:
    st.session_state["cap_val"] = 2.0

topc1, topc2, topc3, topc4 = st.columns([1.1, 1.2, 1.2, 1.5])
with topc1:
    st.session_state["modo_presentacion"] = st.toggle("Modo presentación", value=st.session_state["modo_presentacion"])
with topc2:
    st.session_state["cap_visual"] = st.toggle("Cap visual %", value=st.session_state["cap_visual"])
with topc3:
    cap_options = {"150%": 1.5, "200%": 2.0, "300%": 3.0, "Sin cap": 999.0}
    cap_label = st.selectbox("Máx visual", list(cap_options.keys()), index=1)
    st.session_state["cap_val"] = cap_options[cap_label]
with topc4:
    st.caption("Tip: el cap es solo visual (ranking/gráficos). No altera el cálculo base.")

if st.session_state["modo_presentacion"]:
    st.markdown(hide_sidebar_css(), unsafe_allow_html=True)

# ---------------------------
# FILTROS BASE
# ---------------------------
def apply_obj0_filter(d, show_obj0: bool):
    return d.copy() if show_obj0 else d[d["Obj_val"] > 0].copy()

def apply_cap_visual(d, cap_on: bool, cap_value: float):
    out = d.copy()
    if "Cumpl" not in out.columns:
        return out
    out["Cumpl_plot"] = out["Cumpl"].clip(upper=cap_value) if cap_on else out["Cumpl"]
    return out

meses_disponibles = sorted(df["Mes"].dropna().unique().tolist())
if not meses_disponibles:
    st.error("No se encontraron meses válidos en la base.")
    st.stop()

month_labels = {m: build_month_label(m) for m in meses_disponibles}
meses_labels_ordenados = [month_labels[m] for m in meses_disponibles]
label_to_mes = {v: k for k, v in month_labels.items()}
default_meses = [meses_disponibles[-1]]

def render_filters(area="sidebar"):
    if "meses_sel" not in st.session_state:
        st.session_state["meses_sel"] = default_meses.copy()
    if "semanas_sel" not in st.session_state:
        st.session_state["semanas_sel"] = []
    if "sucursales_sel" not in st.session_state:
        st.session_state["sucursales_sel"] = ["TODAS (Consolidado)"]
    if "show_obj0" not in st.session_state:
        st.session_state["show_obj0"] = True

    container = st.sidebar if area == "sidebar" else st.container()
    sucursales = sorted(df["Sucursal"].dropna().unique().tolist())

    with container:
        if area == "sidebar":
            st.sidebar.markdown("## Filtros obligatorios")
        else:
            st.markdown("### Filtros")

        meses_sel_label = st.multiselect(
            "Mes",
            meses_labels_ordenados,
            default=[month_labels[m] for m in st.session_state["meses_sel"] if m in month_labels],
            key=f"mes_{area}"
        )

        if not meses_sel_label:
            meses_sel_label = [month_labels[default_meses[-1]]]

        st.session_state["meses_sel"] = [label_to_mes[x] for x in meses_sel_label if x in label_to_mes]

        semanas_disponibles = sorted(
            df.loc[df["Mes"].isin(st.session_state["meses_sel"]), "Semana_Num"]
              .dropna().astype(int).unique().tolist()
        )

        if not st.session_state["semanas_sel"]:
            st.session_state["semanas_sel"] = semanas_disponibles.copy()

        st.session_state["semanas_sel"] = [int(x) for x in st.session_state["semanas_sel"] if int(x) in semanas_disponibles]
        if not st.session_state["semanas_sel"]:
            st.session_state["semanas_sel"] = semanas_disponibles.copy()

        semanas_sel = st.multiselect(
            "Semana corte",
            semanas_disponibles,
            default=st.session_state["semanas_sel"],
            key=f"semana_{area}"
        )

        if not semanas_sel:
            semanas_sel = semanas_disponibles.copy()

        st.session_state["semanas_sel"] = [int(x) for x in semanas_sel]

        sucursales_opts = ["TODAS (Consolidado)"] + sucursales
        suc_default = [x for x in st.session_state["sucursales_sel"] if x in sucursales_opts]
        if not suc_default:
            suc_default = ["TODAS (Consolidado)"]

        sucursales_sel = st.multiselect(
            "Sucursal",
            sucursales_opts,
            default=suc_default,
            key=f"sucursal_{area}"
        )

        if not sucursales_sel:
            sucursales_sel = ["TODAS (Consolidado)"]

        if "TODAS (Consolidado)" in sucursales_sel and len(sucursales_sel) > 1:
            sucursales_sel = [x for x in sucursales_sel if x != "TODAS (Consolidado)"]

        st.session_state["sucursales_sel"] = sucursales_sel

        st.markdown("---")
        st.markdown("### Cálculo")
        st.session_state["show_obj0"] = st.checkbox(
            "Incluir filas con Obj=0 (puede distorsionar %)",
            value=st.session_state["show_obj0"],
            key=f"obj0_{area}"
        )

if st.session_state["modo_presentacion"]:
    with st.expander("Abrir filtros (presentación)", expanded=False):
        render_filters(area="top")
else:
    render_filters(area="sidebar")

meses_sel = st.session_state["meses_sel"]
semanas_sel = [int(x) for x in st.session_state["semanas_sel"]]
sucursales_sel = st.session_state["sucursales_sel"]
show_obj0 = bool(st.session_state["show_obj0"])
cap_on = bool(st.session_state["cap_visual"])
cap_val = float(st.session_state["cap_val"])

if not meses_sel:
    meses_sel = default_meses.copy()

if not semanas_sel:
    semanas_sel = sorted(df.loc[df["Mes"].isin(meses_sel), "Semana_Num"].dropna().astype(int).unique().tolist())

if not sucursales_sel:
    sucursales_sel = ["TODAS (Consolidado)"]

selected_all_suc = "TODAS (Consolidado)" in sucursales_sel
selected_sucursales_real = sorted(df["Sucursal"].dropna().unique().tolist()) if selected_all_suc else sucursales_sel

# ---------------------------
# DATOS FILTRADOS
# ---------------------------
df_scope = df[df["Mes"].isin(meses_sel)].copy()

if not selected_all_suc:
    df_scope = df_scope[df_scope["Sucursal"].isin(selected_sucursales_real)].copy()

df_cut = df_scope[df_scope["Semana_Num"].isin(semanas_sel)].copy()
df_month = df_scope.copy()

dias_total_mes = np.nan
dias_transc = np.nan

if meses_sel:
    meses_nums_sel = [int(str(m).split("-")[1]) for m in meses_sel]
    meses_norm_sel = [norm_text(month_name_es(mn)) for mn in meses_nums_sel]
    dias_mes = df_dias_habiles[df_dias_habiles["Mes_norm"].isin(meses_norm_sel)].copy()
    if not dias_mes.empty:
        dias_total_mes = float(dias_mes["Dias habiles"].sum())
        dias_transc = float(dias_mes[dias_mes["Semana"].isin(semanas_sel)]["Dias habiles"].sum())

# ---------------------------
# P&L aperturas del filtro elegido
# ---------------------------
def compute_openings_pl(dfc):
    rep_open = sorted(dfc[(dfc["KPI"].str.upper() == "REPUESTOS") & (dfc["Tipo_KPI"] == "$")]["Categoria_KPI"].dropna().unique().tolist())
    srv_open = sorted(dfc[(dfc["KPI"].str.upper() == "SERVICIOS") & (dfc["Tipo_KPI"] == "$")]["Categoria_KPI"].dropna().unique().tolist())
    return rep_open, srv_open

rep_open, srv_open = compute_openings_pl(df_cut)

if "rep_sel" not in st.session_state:
    st.session_state["rep_sel"] = rep_open
if "srv_sel" not in st.session_state:
    st.session_state["srv_sel"] = srv_open

for x in rep_open:
    if x not in st.session_state["rep_sel"]:
        st.session_state["rep_sel"].append(x)
for x in srv_open:
    if x not in st.session_state["srv_sel"]:
        st.session_state["srv_sel"].append(x)

st.session_state["rep_sel"] = [x for x in st.session_state["rep_sel"] if x in rep_open]
st.session_state["srv_sel"] = [x for x in st.session_state["srv_sel"] if x in srv_open]

def render_pl_multiselect(area="sidebar"):
    container = st.sidebar if area == "sidebar" else st.container()
    with container:
        st.markdown("## Incluir variables (P&L)")
        st.session_state["rep_sel"] = st.multiselect(
            "Repuestos: aperturas incluidas",
            rep_open, default=st.session_state["rep_sel"],
            key=f"rep_sel_{area}"
        )
        st.session_state["srv_sel"] = st.multiselect(
            "Servicios: aperturas incluidas",
            srv_open, default=st.session_state["srv_sel"],
            key=f"srv_sel_{area}"
        )

if st.session_state["modo_presentacion"]:
    with st.expander("Abrir variables P&L (presentación)", expanded=False):
        render_pl_multiselect(area="top")
else:
    st.sidebar.markdown("---")
    render_pl_multiselect(area="sidebar")

rep_sel = st.session_state["rep_sel"]
srv_sel = st.session_state["srv_sel"]

# ---------------------------
# FUNCIONES DE NEGOCIO P&L
# ---------------------------
def summarize_segment(dseg: pd.DataFrame):
    dseg = apply_obj0_filter(dseg, show_obj0)
    real = dseg["Real_val"].sum()
    obj = dseg["Obj_val"].sum()
    c = safe_ratio(real, obj)
    return float(real), float(obj), c

def proyectar_eom_runrate(real_acum: float, dias_trans: float, dias_mes_total: float) -> float:
    if pd.isna(dias_trans) or pd.isna(dias_mes_total) or dias_trans == 0:
        return np.nan
    return (float(real_acum) / float(dias_trans)) * float(dias_mes_total)

def micro_sucursal(d: pd.DataFrame):
    g = d.groupby("Sucursal", as_index=False).agg(Real=("Real_val", "sum"), Obj=("Obj_val", "sum"))
    g["Cumpl"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    g = g[~g["Cumpl"].isna()].copy().sort_values("Cumpl", ascending=False)
    g = apply_cap_visual(g, cap_on, cap_val)
    g["label"] = g.apply(lambda r: f"{pct_str(r['Cumpl'])} | {money_str(r['Real'])}/{money_str(r['Obj'])}", axis=1)
    return g

def micro_aperturas(d: pd.DataFrame):
    g = d.groupby("Categoria_KPI", as_index=False).agg(Real=("Real_val", "sum"), Obj=("Obj_val", "sum"))
    g["Cumpl"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    g = g[~g["Cumpl"].isna()].copy().sort_values("Cumpl", ascending=False)
    g = apply_cap_visual(g, cap_on, cap_val)
    g["label"] = g.apply(lambda r: f"{pct_str(r['Cumpl'])} | {money_str(r['Real'])}/{money_str(r['Obj'])}", axis=1)
    return g

def ranking_sucursal_apertura_micro(d: pd.DataFrame, top_n: int, show_zero: bool):
    x = d.copy()
    if not show_zero:
        x = x[x["Obj_val"] > 0].copy()

    g = x.groupby(["Sucursal", "Categoria_KPI"], as_index=False).agg(
        Real=("Real_val", "sum"),
        Obj=("Obj_val", "sum")
    )
    g["Cumpl"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    g = g[~g["Cumpl"].isna()].copy().sort_values("Cumpl", ascending=False).head(top_n).copy()
    g = apply_cap_visual(g, cap_on, cap_val)
    g["label"] = g.apply(lambda r: f"{pct_str(r['Cumpl'])} | {money_str(r['Real'])}/{money_str(r['Obj'])}", axis=1)
    g["key"] = g["Sucursal"].astype(str) + " — " + g["Categoria_KPI"].astype(str)
    return g

def principal_driver_gap(d_pl: pd.DataFrame):
    x = apply_obj0_filter(d_pl.copy(), show_obj0)
    if x.empty:
        return None
    g = x.groupby(["KPI", "Categoria_KPI"], as_index=False).agg(
        Real=("Real_val", "sum"),
        Obj=("Obj_val", "sum")
    )
    g["Gap"] = g["Obj"] - g["Real"]
    g = g.sort_values("Gap", ascending=False)
    row = g.iloc[0]
    return {"KPI": str(row["KPI"]), "Cat": str(row["Categoria_KPI"]), "Gap": float(row["Gap"])}

def spark_evolucion(df_scope_month: pd.DataFrame, chart_key: str):
    if df_scope_month is None or df_scope_month.empty:
        return

    x = apply_obj0_filter(df_scope_month.copy(), show_obj0)
    g = (
        x.groupby(["Mes", "Semana_Mes"], as_index=False)
         .agg(Real=("Real_val", "sum"), Obj=("Obj_val", "sum"))
         .sort_values(["Mes", "Semana_Mes"])
    )
    g["Cumpl_sem"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    g = g[~g["Cumpl_sem"].isna()].copy()

    if g.empty:
        return

    g["Mes_Label"] = g["Mes"].apply(build_month_label)
    g["Eje"] = g["Mes_Label"] + " - S" + g["Semana_Mes"].astype(str)

    gg = g.copy()
    gg["Cumpl"] = gg["Cumpl_sem"]
    gg["Cumpl_plot"] = gg["Cumpl"].clip(upper=cap_val) if cap_on else gg["Cumpl"]
    gg["txt"] = gg["Cumpl"].apply(pct_str)

    fig = px.line(gg, x="Eje", y="Cumpl_plot", markers=True, text="txt")
    fig.update_traces(mode="lines+markers+text", textposition="top center")
    fig.update_layout(
        height=150,
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis_title="",
        yaxis_title=""
    )
    fig.update_yaxes(tickformat=".0%")

    st.plotly_chart(fig, use_container_width=True, key=chart_key)

def filter_hitos_by_month(df_hitos: pd.DataFrame, meses_keys: list[str]) -> pd.DataFrame:
    if df_hitos is None or df_hitos.empty:
        return pd.DataFrame()

    tmp = df_hitos.copy()
    month_cols = []
    for c in tmp.columns:
        nc = norm_text(c)
        if nc in {"mes", "periodo", "periodo mes", "fecha", "month"} or "mes" in nc or "periodo" in nc:
            month_cols.append(c)

    if not month_cols:
        return tmp

    mask = pd.Series(False, index=tmp.index)

    for mes_key in meses_keys:
        mes_label = build_month_label(mes_key)
        mes_norm = norm_text(mes_label)
        mes_key_norm = norm_text(mes_key)
        nombre_mes = norm_text(month_name_es(int(str(mes_key).split("-")[1])))

        for c in month_cols:
            s = tmp[c].astype(str).apply(norm_text)
            mask = (
                mask
                | s.str.contains(mes_key_norm, na=False)
                | s.str.contains(nombre_mes, na=False)
                | s.str.contains(mes_norm, na=False)
            )

    filtrado = tmp[mask].copy()
    return filtrado if not filtrado.empty else tmp

# ---------------------------
# HEADER + EXCEL PROFESIONAL
# ---------------------------
st.title("Tablero Posventa — Macro → Micro (Semanal + Acumulado)")

hl, hr = st.columns([2.2, 1.0])
with hl:
    st.caption(
        f"Mes(es): **{labels_from_mes_keys(meses_sel)}** | "
        f"Semana(s): **{list_to_export_text(semanas_sel)}** | "
        f"Sucursal(es): **{('TODAS (Consolidado)' if selected_all_suc else list_to_export_text(selected_sucursales_real))}** | "
        f"Días hábiles transcurridos: **{(int(dias_transc) if not pd.isna(dias_transc) else '—')}** | "
        f"Días hábiles período: **{(int(dias_total_mes) if not pd.isna(dias_total_mes) else '—')}**"
    )

with hr:
    semanas_export_opts = sorted(df.loc[df["Mes"].isin(meses_sel), "Semana_Num"].dropna().astype(int).unique().tolist())
    sem_export = st.multiselect(
        "Semana(s) para Excel",
        semanas_export_opts,
        default=semanas_sel,
        help="El Excel se calcula respetando estas semanas dentro de los meses seleccionados."
    )
    if not sem_export:
        sem_export = semanas_export_opts.copy()

df_cut_xls = df[df["Mes"].isin(meses_sel)].copy()
if not selected_all_suc:
    df_cut_xls = df_cut_xls[df_cut_xls["Sucursal"].isin(selected_sucursales_real)].copy()
df_cut_xls = df_cut_xls[df_cut_xls["Semana_Num"].isin([int(x) for x in sem_export])].copy()

dias_total_mes_xls = np.nan
dias_transc_xls = np.nan

if meses_sel:
    meses_nums_sel_xls = [int(str(m).split("-")[1]) for m in meses_sel]
    meses_norm_sel_xls = [norm_text(month_name_es(mn)) for mn in meses_nums_sel_xls]
    dias_mes_xls = df_dias_habiles[df_dias_habiles["Mes_norm"].isin(meses_norm_sel_xls)].copy()
    if not dias_mes_xls.empty:
        dias_total_mes_xls = float(dias_mes_xls["Dias habiles"].sum())
        dias_transc_xls = float(dias_mes_xls[dias_mes_xls["Semana"].isin([int(x) for x in sem_export])]["Dias habiles"].sum())

d_pl_xls = df_cut_xls[df_cut_xls["Tipo_KPI"] == "$"].copy()
d_rep_xls = d_pl_xls[d_pl_xls["KPI"].str.upper() == "REPUESTOS"].copy()
d_rep_xls = d_rep_xls[d_rep_xls["Categoria_KPI"].isin(rep_sel)].copy()
d_srv_xls = d_pl_xls[d_pl_xls["KPI"].str.upper() == "SERVICIOS"].copy()
d_srv_xls = d_srv_xls[d_srv_xls["Categoria_KPI"].isin(srv_sel)].copy()

rep_real_xls, rep_obj_xls, rep_c_xls = summarize_segment(d_rep_xls)
srv_real_xls, srv_obj_xls, srv_c_xls = summarize_segment(d_srv_xls)

total_real_xls = rep_real_xls + srv_real_xls
total_obj_xls = rep_obj_xls + srv_obj_xls
total_c_xls = safe_ratio(total_real_xls, total_obj_xls)

rep_proy_xls = proyectar_eom_runrate(rep_real_xls, dias_transc_xls, dias_total_mes_xls)
srv_proy_xls = proyectar_eom_runrate(srv_real_xls, dias_transc_xls, dias_total_mes_xls)
total_proy_xls = proyectar_eom_runrate(total_real_xls, dias_transc_xls, dias_total_mes_xls)

rep_by_suc_xls = micro_sucursal(apply_obj0_filter(d_rep_xls, show_obj0))
srv_by_suc_xls = micro_sucursal(apply_obj0_filter(d_srv_xls, show_obj0))
rep_by_ap_xls = micro_aperturas(apply_obj0_filter(d_rep_xls, show_obj0))
srv_by_ap_xls = micro_aperturas(apply_obj0_filter(d_srv_xls, show_obj0))

top_n_xls = int(st.session_state.get("top_n_rank", 10))
show_zero_rank_xls = bool(st.session_state.get("show_zero_rank", False))
rep_rank_xls = ranking_sucursal_apertura_micro(d_rep_xls, top_n=top_n_xls, show_zero=show_zero_rank_xls)
srv_rank_xls = ranking_sucursal_apertura_micro(d_srv_xls, top_n=top_n_xls, show_zero=show_zero_rank_xls)

driver_xls = principal_driver_gap(pd.concat([d_rep_xls, d_srv_xls], ignore_index=True))
hitos_export = filter_hitos_by_month(df_resumen_mes, meses_sel)

abiertas_std = build_operational_standard(df_abiertas_raw, "Abiertas")
pfact_std = build_operational_standard(df_pfact_raw, "Pendientes Fact")
presup_std = build_operational_standard(df_presup_raw, "Presupuestos")

resumen_xls = pd.DataFrame([
    {
        "Bloque": "Repuestos", "Real_Acum": rep_real_xls, "Obj_Acum": rep_obj_xls, "Cumpl_Acum": rep_c_xls,
        "Proy_EOM_RunRate": rep_proy_xls, "Dias_Transc": (np.nan if pd.isna(dias_transc_xls) else dias_transc_xls),
        "Dias_Mes": (np.nan if pd.isna(dias_total_mes_xls) else dias_total_mes_xls)
    },
    {
        "Bloque": "Servicios", "Real_Acum": srv_real_xls, "Obj_Acum": srv_obj_xls, "Cumpl_Acum": srv_c_xls,
        "Proy_EOM_RunRate": srv_proy_xls, "Dias_Transc": (np.nan if pd.isna(dias_transc_xls) else dias_transc_xls),
        "Dias_Mes": (np.nan if pd.isna(dias_total_mes_xls) else dias_total_mes_xls)
    },
    {
        "Bloque": "Total Postventa", "Real_Acum": total_real_xls, "Obj_Acum": total_obj_xls, "Cumpl_Acum": total_c_xls,
        "Proy_EOM_RunRate": total_proy_xls, "Dias_Transc": (np.nan if pd.isna(dias_transc_xls) else dias_transc_xls),
        "Dias_Mes": (np.nan if pd.isna(dias_total_mes_xls) else dias_total_mes_xls)
    },
])

meta_xls = pd.DataFrame([{
    "Meses_Seleccionados": labels_from_mes_keys(meses_sel),
    "Meses_Key": list_to_export_text(meses_sel),
    "Semanas_Dashboard": list_to_export_text(semanas_sel),
    "Semanas_Export_Excel": list_to_export_text([int(x) for x in sem_export]),
    "Sucursales_Seleccionadas": ("TODAS (Consolidado)" if selected_all_suc else list_to_export_text(selected_sucursales_real)),
    "Dias_Habiles_Transcurridos_Excel": (None if pd.isna(dias_transc_xls) else int(dias_transc_xls)),
    "Dias_Habiles_Periodo_Excel": (None if pd.isna(dias_total_mes_xls) else int(dias_total_mes_xls)),
    "Cap_Visual_On": cap_on,
    "Cap_Visual_Max": cap_val,
    "Incluir_Obj_0": show_obj0,
    "Hoja_Dias_Habiles": (DIAS_SHEET_FOUND if DIAS_SHEET_FOUND is not None else "NO_ENCONTRADA"),
    "Hoja_Resumen_del_Mes": (RESUMEN_SHEET_FOUND if RESUMEN_SHEET_FOUND is not None else "NO_ENCONTRADA"),
    "Hoja_Abiertas": (ABIERTAS_SHEET_FOUND if ABIERTAS_SHEET_FOUND is not None else "NO_ENCONTRADA"),
    "Hoja_Pend_Fact": (PFACT_SHEET_FOUND if PFACT_SHEET_FOUND is not None else "NO_ENCONTRADA"),
    "Hoja_Presupuestos": (PRESUP_SHEET_FOUND if PRESUP_SHEET_FOUND is not None else "NO_ENCONTRADA"),
    "Ranking_TopN": top_n_xls,
    "Ranking_ShowZero": show_zero_rank_xls,
    "Principal_Desvio_KPI": (driver_xls["KPI"] if driver_xls else "—"),
    "Principal_Desvio_Cat": (driver_xls["Cat"] if driver_xls else "—"),
    "Principal_Desvio_Gap": (driver_xls["Gap"] if driver_xls else np.nan),
}])

excel_bytes_prof = build_exec_excel_professional(
    meta_df=meta_xls,
    resumen_df=resumen_xls,
    pl_sucursal_rep=rep_by_suc_xls,
    pl_sucursal_srv=srv_by_suc_xls,
    pl_ap_rep=rep_by_ap_xls,
    pl_ap_srv=srv_by_ap_xls,
    ranking_rep=rep_rank_xls,
    ranking_srv=srv_rank_xls,
    hitos_df=hitos_export,
    abiertas_df=abiertas_std,
    pendientes_fact_df=pfact_std,
    presup_df=presup_std,
)

with hr:
    meses_file = "_".join([str(m).replace("-", "") for m in meses_sel]) if meses_sel else "sin_mes"
    semanas_file = "_".join([str(int(x)) for x in sem_export]) if sem_export else "sin_sem"
    suc_file = "TODAS" if selected_all_suc else "_".join([str(s).replace(" ", "_") for s in selected_sucursales_real])

    st.download_button(
        "⬇️ Resumen Ejecutivo (Excel)",
        data=excel_bytes_prof,
        file_name=f"Resumen_Ejecutivo_Tablero_Posventa_{meses_file}_Sem_{semanas_file}_{suc_file}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )



# ============================================================
# TAB NUEVO — CIERRE DE MES / GAP A OBJETIVO
# ============================================================
def _fmt_by_tipo(valor, tipo_kpi):
    """Formatea $ como dinero y Q como cantidad, respetando la lógica existente del tablero."""
    if str(tipo_kpi).strip() == "$":
        return money_str(valor)
    return qty_str(valor)

def _estado_gap(cumpl):
    if pd.isna(cumpl):
        return "⚪ Sin objetivo"
    if cumpl >= 1:
        return "🟢 Cumplido"
    if cumpl >= 0.90:
        return "🟡 Cerca"
    return "🔴 En riesgo"

def _bloque_gap(row):
    kpi = str(row.get("KPI", "")).upper()
    cat = str(row.get("Categoria_KPI", "")).upper()
    if kpi == "REPUESTOS":
        return "Repuestos"
    if kpi == "SERVICIOS":
        return "Servicios"
    if "CPU" in kpi or "CPU" in cat:
        return "CPUs"
    if "NEUM" in kpi or "NEUM" in cat:
        return "Neumáticos"
    return str(row.get("KPI", "Otros"))

def render_tab_cierre_gap():
    st.markdown("## 🎯 Cierre de Mes — GAP a Objetivo")
    st.caption(
        "Vista de gestión para la última semana: cuánto falta, por sucursal y por variable, "
        "respetando mes, semana(s), sucursal(es) y aperturas P&L seleccionadas."
    )
    st.markdown("---")

    base = df_cut.copy()
    if base.empty:
        st.info("No hay datos para los filtros seleccionados.")
        return

    rep_options = rep_open.copy()
    srv_options = srv_open.copy()

    rep_default = [x for x in rep_sel if x in rep_options]
    srv_default = [x for x in srv_sel if x in srv_options]

    c1, c2 = st.columns(2)
    with c1:
        rep_gap_sel = st.multiselect(
            "Repuestos: aperturas incluidas",
            rep_options,
            default=rep_default,
            key="gap_cierre_rep_sel"
        )
    with c2:
        srv_gap_sel = st.multiselect(
            "Servicios: aperturas incluidas",
            srv_options,
            default=srv_default,
            key="gap_cierre_srv_sel"
        )

    c3, c4, c5 = st.columns([1.0, 1.0, 1.3])
    with c3:
        incluir_cpus = st.checkbox("Incluir CPUs", value=True, key="gap_incluir_cpus")
    with c4:
        incluir_neum = st.checkbox("Incluir Neumáticos", value=True, key="gap_incluir_neum")
    with c5:
        solo_pendientes = st.checkbox("🔥 Modo cierre: ver solo pendientes", value=False, key="gap_solo_pendientes")

    partes = []

    d_rep_gap = base[
        (base["Tipo_KPI"] == "$") &
        (base["KPI"].str.upper() == "REPUESTOS") &
        (base["Categoria_KPI"].isin(rep_gap_sel))
    ].copy()
    if not d_rep_gap.empty:
        partes.append(d_rep_gap)

    d_srv_gap = base[
        (base["Tipo_KPI"] == "$") &
        (base["KPI"].str.upper() == "SERVICIOS") &
        (base["Categoria_KPI"].isin(srv_gap_sel))
    ].copy()
    if not d_srv_gap.empty:
        partes.append(d_srv_gap)

    if incluir_cpus:
        d_cpu = base[
            base["KPI"].astype(str).str.contains("CPU", case=False, na=False) |
            base["Categoria_KPI"].astype(str).str.contains("CPU", case=False, na=False)
        ].copy()
        if not d_cpu.empty:
            partes.append(d_cpu)

    if incluir_neum:
        d_neum = base[
            base["KPI"].astype(str).str.contains("NEUM", case=False, na=False) |
            base["Categoria_KPI"].astype(str).str.contains("NEUM", case=False, na=False)
        ].copy()
        if not d_neum.empty:
            partes.append(d_neum)

    if not partes:
        st.warning("No hay datos para las variables seleccionadas.")
        return

    work = pd.concat(partes, ignore_index=True).drop_duplicates()
    work = apply_obj0_filter(work, show_obj0)

    if work.empty:
        st.warning("No hay datos luego de aplicar el filtro Obj=0.")
        return

    work["Bloque"] = work.apply(_bloque_gap, axis=1)

    gap = (
        work.groupby(["Sucursal", "Bloque", "KPI", "Categoria_KPI", "Tipo_KPI"], as_index=False)
        .agg(Real=("Real_val", "sum"), Obj=("Obj_val", "sum"))
    )
    gap["GAP"] = gap["Real"] - gap["Obj"]
    gap["Falta"] = np.where(gap["GAP"] < 0, gap["GAP"].abs(), 0)
    gap["Cumpl"] = gap.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    gap["Estado"] = gap["Cumpl"].apply(_estado_gap)

    dias_restantes = np.nan
    if not pd.isna(dias_total_mes) and not pd.isna(dias_transc):
        dias_restantes = max(float(dias_total_mes) - float(dias_transc), 0)

    if not pd.isna(dias_restantes) and dias_restantes > 0:
        gap["Necesario_Dia"] = np.where(gap["Falta"] > 0, gap["Falta"] / dias_restantes, np.nan)
    else:
        gap["Necesario_Dia"] = np.nan

    if solo_pendientes:
        gap = gap[gap["Falta"] > 0].copy()

    if gap.empty:
        st.success("Con los filtros seleccionados no hay pendientes: todo está cumplido o sin falta.")
        return

    total_real = gap["Real"].sum()
    total_obj = gap["Obj"].sum()
    total_cumpl = safe_ratio(total_real, total_obj)
    total_gap = total_real - total_obj
    total_falta = gap["Falta"].sum()
    suc_pend = gap[gap["Falta"] > 0]["Sucursal"].nunique()

    st.markdown("### 📌 Foto ejecutiva del cierre")
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(card_html_base("Cumplimiento total", pct_str(total_cumpl), "Real / Objetivo acumulado"), unsafe_allow_html=True)
    with k2:
        st.markdown(card_html_base("GAP total", money_str(total_gap), "Real - Objetivo"), unsafe_allow_html=True)
    with k3:
        st.markdown(card_html_base("Falta total", money_str(total_falta), "Solo pendientes negativos"), unsafe_allow_html=True)
    with k4:
        st.markdown(card_html_base("Sucursales con pendiente", qty_str(suc_pend), "Con al menos una variable bajo objetivo"), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📍 GAP acumulado por sucursal")

    suc = (
        gap.groupby("Sucursal", as_index=False)
        .agg(Real=("Real", "sum"), Obj=("Obj", "sum"), Falta=("Falta", "sum"))
    )
    suc["GAP"] = suc["Real"] - suc["Obj"]
    suc["Cumpl"] = suc.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    suc["label"] = suc.apply(lambda r: f"{money_str(r['GAP'])} | {pct_str(r['Cumpl'])}", axis=1)
    suc = suc.sort_values("GAP", ascending=True)

    fig = px.bar(
        suc,
        x="GAP",
        y="Sucursal",
        orientation="h",
        text="label",
        title="GAP vs objetivo acumulado"
    )
    fig.add_vline(x=0, line_width=2, line_dash="dash")
    fig.update_layout(height=420, margin=dict(l=10, r=10, t=45, b=10), xaxis_title="GAP (Real - Objetivo)")
    fig.update_traces(textposition="inside")
    st.plotly_chart(fig, use_container_width=True, key="gap_cierre_sucursal")

    st.markdown("### 🚦 Matriz de foco — Sucursal x variable")

    tabla = gap.copy().sort_values(["Falta", "Cumpl"], ascending=[False, True])
    tabla["Variable"] = np.where(
        tabla["KPI"].str.upper().isin(["REPUESTOS", "SERVICIOS"]),
        tabla["Categoria_KPI"],
        tabla["KPI"]
    )

    tabla_show = tabla[[
        "Sucursal", "Bloque", "Variable", "Tipo_KPI", "Real", "Obj", "Cumpl", "GAP", "Falta", "Necesario_Dia", "Estado"
    ]].copy()

    tabla_show["Real"] = tabla_show.apply(lambda r: _fmt_by_tipo(r["Real"], r["Tipo_KPI"]), axis=1)
    tabla_show["Objetivo"] = tabla.apply(lambda r: _fmt_by_tipo(r["Obj"], r["Tipo_KPI"]), axis=1)
    tabla_show["Cumplimiento"] = tabla_show["Cumpl"].apply(pct_str)
    tabla_show["GAP"] = tabla.apply(lambda r: _fmt_by_tipo(r["GAP"], r["Tipo_KPI"]), axis=1)
    tabla_show["Falta"] = tabla.apply(lambda r: _fmt_by_tipo(r["Falta"], r["Tipo_KPI"]), axis=1)
    tabla_show["Necesario por día"] = tabla.apply(lambda r: _fmt_by_tipo(r["Necesario_Dia"], r["Tipo_KPI"]) if pd.notna(r["Necesario_Dia"]) else "—", axis=1)

    tabla_show = tabla_show[[
        "Sucursal", "Bloque", "Variable", "Real", "Objetivo", "Cumplimiento", "GAP", "Falta", "Necesario por día", "Estado"
    ]]

    st.dataframe(tabla_show, use_container_width=True, hide_index=True)

    st.markdown("### ⚔️ Prioridad de acción")
    pendientes = tabla[tabla["Falta"] > 0].copy().sort_values("Falta", ascending=False).head(15)

    if pendientes.empty:
        st.success("No hay pendientes para recuperar con los filtros seleccionados.")
    else:
        pendientes["Variable"] = np.where(
            pendientes["KPI"].str.upper().isin(["REPUESTOS", "SERVICIOS"]),
            pendientes["Categoria_KPI"],
            pendientes["KPI"]
        )
        pendientes["Eje"] = pendientes["Sucursal"].astype(str) + " — " + pendientes["Variable"].astype(str)
        pendientes["label"] = pendientes.apply(lambda r: _fmt_by_tipo(r["Falta"], r["Tipo_KPI"]), axis=1)

        fig2 = px.bar(
            pendientes.sort_values("Falta", ascending=True),
            x="Falta",
            y="Eje",
            orientation="h",
            text="label",
            title="Top pendientes a recuperar"
        )
        fig2.update_layout(height=520, margin=dict(l=10, r=10, t=45, b=10), xaxis_title="Falta para cumplir")
        fig2.update_traces(textposition="inside")
        st.plotly_chart(fig2, use_container_width=True, key="gap_cierre_prioridad")

        st.info(
            "Lectura sugerida para la reunión: empezar por los mayores pendientes absolutos, "
            "validar si todavía son recuperables esta semana y asignar responsable por sucursal/variable."
        )



# ============================================================
# NUEVO NIVEL — DIRECCIÓN + GESTIÓN INTELIGENTE
# ============================================================
def _status_from_cumpl(cumpl):
    if pd.isna(cumpl):
        return {
            "emoji": "⚪",
            "label": "Sin lectura",
            "msg": "No hay objetivo suficiente para calcular cumplimiento.",
            "color": "gray"
        }
    if cumpl >= 1:
        return {
            "emoji": "🟢",
            "label": "En objetivo",
            "msg": "Postventa está cumpliendo o superando el objetivo acumulado.",
            "color": "green"
        }
    if cumpl >= 0.90:
        return {
            "emoji": "🟡",
            "label": "Cerca del objetivo",
            "msg": "Postventa está cerca del objetivo, pero requiere seguimiento de desvíos.",
            "color": "yellow"
        }
    return {
        "emoji": "🔴",
        "label": "En riesgo",
        "msg": "Postventa presenta un desvío relevante frente al objetivo acumulado.",
        "color": "red"
    }

def _fmt_gap(valor, tipo_kpi):
    if str(tipo_kpi).strip() == "$":
        return money_str(valor)
    return qty_str(valor)

def _metric_delta_pts(cumpl):
    if pd.isna(cumpl):
        return None
    return f"{(float(cumpl) - 1) * 100:.1f} pts vs objetivo"

def build_direction_context():
    d_pl = df_cut[df_cut["Tipo_KPI"] == "$"].copy()

    d_rep = d_pl[d_pl["KPI"].str.upper() == "REPUESTOS"].copy()
    d_rep = d_rep[d_rep["Categoria_KPI"].isin(rep_sel)].copy()

    d_srv = d_pl[d_pl["KPI"].str.upper() == "SERVICIOS"].copy()
    d_srv = d_srv[d_srv["Categoria_KPI"].isin(srv_sel)].copy()

    rep_real, rep_obj, rep_c = summarize_segment(d_rep)
    srv_real, srv_obj, srv_c = summarize_segment(d_srv)

    total_real = rep_real + srv_real
    total_obj = rep_obj + srv_obj
    total_c = safe_ratio(total_real, total_obj)
    total_proy = proyectar_eom_runrate(total_real, dias_transc, dias_total_mes)

    gap_total = total_real - total_obj
    falta_total = max(total_obj - total_real, 0)

    d_q = apply_obj0_filter(df_cut[df_cut["Tipo_KPI"] != "$"].copy(), show_obj0)
    q_real = d_q["Real_val"].sum() if not d_q.empty else np.nan
    q_obj = d_q["Obj_val"].sum() if not d_q.empty else np.nan
    q_c = safe_ratio(q_real, q_obj) if not d_q.empty else np.nan

    driver = principal_driver_gap(pd.concat([d_rep, d_srv], ignore_index=True))

    op_ab = op_summary(abiertas_std) if "abiertas_std" in globals() else {"count": 0, "monto": np.nan, "age_avg": np.nan, "age_max": np.nan}
    op_pf = op_summary(pfact_std) if "pfact_std" in globals() else {"count": 0, "monto": np.nan, "age_avg": np.nan, "age_max": np.nan}
    op_pr = op_summary(presup_std) if "presup_std" in globals() else {"count": 0, "monto": np.nan, "age_avg": np.nan, "age_max": np.nan}

    return {
        "rep_real": rep_real, "rep_obj": rep_obj, "rep_c": rep_c,
        "srv_real": srv_real, "srv_obj": srv_obj, "srv_c": srv_c,
        "total_real": total_real, "total_obj": total_obj, "total_c": total_c,
        "total_proy": total_proy, "gap_total": gap_total, "falta_total": falta_total,
        "q_real": q_real, "q_obj": q_obj, "q_c": q_c,
        "driver": driver,
        "d_rep": d_rep, "d_srv": d_srv,
        "op_abiertas": op_ab, "op_pfact": op_pf, "op_presup": op_pr
    }

def build_direction_narrative(ctx):
    status = _status_from_cumpl(ctx["total_c"])
    partes = [f"{status['emoji']} {status['msg']}"]

    if pd.notna(ctx["total_c"]):
        partes.append(f"Cumplimiento acumulado: {pct_str(ctx['total_c'])}.")

    if ctx["falta_total"] > 0:
        partes.append(f"Falta recuperar {money_str(ctx['falta_total'])} para llegar al objetivo del corte.")

    if pd.notna(ctx["total_proy"]) and pd.notna(ctx["total_obj"]) and ctx["total_obj"] > 0:
        proy_c = safe_ratio(ctx["total_proy"], ctx["total_obj"])
        if pd.notna(proy_c):
            if proy_c >= 1:
                partes.append(f"Con el ritmo actual, la proyección de cierre queda en {pct_str(proy_c)} del objetivo.")
            else:
                partes.append(f"Con el ritmo actual, la proyección de cierre quedaría en {pct_str(proy_c)} del objetivo.")

    if ctx["driver"]:
        partes.append(f"El principal foco económico es {ctx['driver']['KPI']} / {ctx['driver']['Cat']} con gap de {money_str(ctx['driver']['Gap'])}.")

    if pd.notna(ctx["q_c"]) and ctx["q_c"] < 1:
        partes.append(f"Los KPIs de volumen/cantidad están al {pct_str(ctx['q_c'])}; revisar tráfico, conversión y productividad.")

    return " ".join(partes)

def build_action_recommendations(ctx):
    acciones = []

    if pd.notna(ctx["total_c"]) and ctx["total_c"] < 0.90:
        acciones.append("Priorizar recuperación del gap económico por sucursal y apertura antes de profundizar análisis secundarios.")
    elif pd.notna(ctx["total_c"]) and ctx["total_c"] < 1:
        acciones.append("Mantener control diario del avance, porque el desvío todavía parece recuperable.")
    else:
        acciones.append("Sostener ritmo y proteger margen; evitar que el cumplimiento se explique solo por precio o mix.")

    if ctx["driver"]:
        acciones.append(f"Asignar responsable específico para {ctx['driver']['KPI']} / {ctx['driver']['Cat']} y revisar plan de cierre.")

    if ctx["op_pfact"]["monto"] is not None and pd.notna(ctx["op_pfact"]["monto"]) and ctx["op_pfact"]["monto"] > 0:
        acciones.append(f"Atacar pendientes de facturación: hay {money_str(ctx['op_pfact']['monto'])} de potencial administrativo.")

    if ctx["op_abiertas"]["count"] > 0:
        acciones.append(f"Revisar órdenes abiertas: {qty_str(ctx['op_abiertas']['count'])} casos activos pueden convertirse en facturación o demora operativa.")

    if ctx["op_presup"]["count"] > 0:
        acciones.append(f"Activar seguimiento comercial de presupuestos: {qty_str(ctx['op_presup']['count'])} oportunidades pendientes.")

    return acciones[:5]

def render_direction_heatmap(ctx):
    base = pd.concat([ctx["d_rep"], ctx["d_srv"]], ignore_index=True)
    base = apply_obj0_filter(base, show_obj0)

    if base.empty:
        st.info("No hay datos económicos para construir la matriz Dirección.")
        return

    g = base.groupby(["Sucursal", "KPI"], as_index=False).agg(
        Real=("Real_val", "sum"),
        Obj=("Obj_val", "sum")
    )
    g["Cumpl"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    g["GAP"] = g["Real"] - g["Obj"]
    g = g[~g["Cumpl"].isna()].copy()

    if g.empty:
        st.info("No hay cumplimiento calculable para la matriz.")
        return

    g["Estado"] = g["Cumpl"].apply(lambda x: _status_from_cumpl(x)["label"])
    g_show = g.sort_values(["Cumpl", "GAP"], ascending=[True, True]).copy()
    g_show["Cumplimiento"] = g_show["Cumpl"].apply(pct_str)
    g_show["Real"] = g_show["Real"].apply(money_str)
    g_show["Objetivo"] = g_show["Obj"].apply(money_str)
    g_show["GAP"] = g_show["GAP"].apply(money_str)

    st.dataframe(
        g_show[["Sucursal", "KPI", "Real", "Objetivo", "Cumplimiento", "GAP", "Estado"]],
        use_container_width=True,
        hide_index=True
    )

def render_direction_tab():
    st.markdown("## 🏢 Dirección — Resumen Ejecutivo")
    st.caption(
        "Vista de lectura rápida: resultado, proyección, causa principal y focos de acción. "
        "Respeta los filtros de mes, semana, sucursal y aperturas P&L."
    )
    st.markdown("---")

    ctx = build_direction_context()
    status = _status_from_cumpl(ctx["total_c"])

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Facturación acumulada", money_str(ctx["total_real"]), delta=_metric_delta_pts(ctx["total_c"]))
    with k2:
        st.metric("Objetivo acumulado", money_str(ctx["total_obj"]))
    with k3:
        st.metric("Cumplimiento", pct_str(ctx["total_c"]), delta=_metric_delta_pts(ctx["total_c"]))
    with k4:
        st.metric("Proyección cierre", money_str(ctx["total_proy"]))

    if status["color"] == "green":
        st.success(f"{status['emoji']} {status['label']} — {status['msg']}")
    elif status["color"] == "yellow":
        st.warning(f"{status['emoji']} {status['label']} — {status['msg']}")
    elif status["color"] == "red":
        st.error(f"{status['emoji']} {status['label']} — {status['msg']}")
    else:
        st.info(f"{status['emoji']} {status['label']} — {status['msg']}")

    st.markdown("### 🧠 Narrativa automática para Dirección")
    st.info(build_direction_narrative(ctx))

    st.markdown("### 🔥 Focos de acción sugeridos")
    acciones = build_action_recommendations(ctx)
    for i, acc in enumerate(acciones, start=1):
        st.markdown(f"**{i}.** {acc}")

    st.markdown("---")
    st.markdown("### 🧩 Lectura por bloque")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(card_html_base(
            "Repuestos",
            pct_str(ctx["rep_c"]),
            f"Real {money_str(ctx['rep_real'])} | Obj {money_str(ctx['rep_obj'])}"
        ), unsafe_allow_html=True)
    with c2:
        st.markdown(card_html_base(
            "Servicios",
            pct_str(ctx["srv_c"]),
            f"Real {money_str(ctx['srv_real'])} | Obj {money_str(ctx['srv_obj'])}"
        ), unsafe_allow_html=True)
    with c3:
        st.markdown(card_html_base(
            "Volumen / Q",
            pct_str(ctx["q_c"]),
            f"Real {qty_str(ctx['q_real'])} | Obj {qty_str(ctx['q_obj'])}"
        ), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### ⚙️ Riesgo operativo que puede convertirse en facturación")

    o1, o2, o3 = st.columns(3)
    with o1:
        st.markdown(card_html_base(
            "Órdenes abiertas",
            qty_str(ctx["op_abiertas"]["count"]),
            f"Potencial {money_str(ctx['op_abiertas']['monto'])} | Antig. máx {qty_str(ctx['op_abiertas']['age_max'])} días"
        ), unsafe_allow_html=True)
    with o2:
        st.markdown(card_html_base(
            "Pend. facturación",
            qty_str(ctx["op_pfact"]["count"]),
            f"Potencial {money_str(ctx['op_pfact']['monto'])} | Antig. máx {qty_str(ctx['op_pfact']['age_max'])} días"
        ), unsafe_allow_html=True)
    with o3:
        st.markdown(card_html_base(
            "Presupuestos",
            qty_str(ctx["op_presup"]["count"]),
            f"Potencial {money_str(ctx['op_presup']['monto'])} | Antig. máx {qty_str(ctx['op_presup']['age_max'])} días"
        ), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📊 Evolución semanal del cumplimiento económico")
    total_month_dir = pd.concat([
        df_month[(df_month["Tipo_KPI"] == "$") & (df_month["KPI"].str.upper() == "REPUESTOS") & (df_month["Categoria_KPI"].isin(rep_sel))],
        df_month[(df_month["Tipo_KPI"] == "$") & (df_month["KPI"].str.upper() == "SERVICIOS") & (df_month["Categoria_KPI"].isin(srv_sel))]
    ], ignore_index=True)
    spark_evolucion(total_month_dir, chart_key="spark_direccion_total")

    st.markdown("---")
    st.markdown("### 🚦 Matriz Dirección — Sucursal x bloque")
    render_direction_heatmap(ctx)

    st.markdown("---")
    with st.expander("🧾 Texto listo para copiar al mail / comité", expanded=False):
        st.write(build_direction_narrative(ctx))
        st.write("Acciones sugeridas:")
        for i, acc in enumerate(acciones, start=1):
            st.write(f"{i}. {acc}")

# ---------------------------
# TABS
# ---------------------------
tab0, tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "🏢 Dirección",
    "🧩 P&L (Repuestos vs Servicios)",
    "📌 KPIs (resto)",
    "🧪 Gestión (desvíos)",
    "🗓️ Hitos del mes",
    "🔧 Órdenes Abiertas",
    "🧾 Pend. Facturación",
    "💬 Presupuestos",
    "🎯 Cierre GAP"
])
# ============================================================
# TAB 0 — DIRECCIÓN
# ============================================================
with tab0:
    render_direction_tab()

# ============================================================
# TAB 1 — P&L
# ============================================================
with tab1:
    st.markdown("## 🧩 P&L — Macro → Micro")
    st.markdown("---")

    d_pl = df_cut[df_cut["Tipo_KPI"] == "$"].copy()

    d_rep = d_pl[d_pl["KPI"].str.upper() == "REPUESTOS"].copy()
    d_rep = d_rep[d_rep["Categoria_KPI"].isin(rep_sel)].copy()

    d_srv = d_pl[d_pl["KPI"].str.upper() == "SERVICIOS"].copy()
    d_srv = d_srv[d_srv["Categoria_KPI"].isin(srv_sel)].copy()

    rep_real, rep_obj, rep_c = summarize_segment(d_rep)
    srv_real, srv_obj, srv_c = summarize_segment(d_srv)

    total_real = rep_real + srv_real
    total_obj = rep_obj + srv_obj
    total_c = safe_ratio(total_real, total_obj)

    rep_proy = proyectar_eom_runrate(rep_real, dias_transc, dias_total_mes)
    srv_proy = proyectar_eom_runrate(srv_real, dias_transc, dias_total_mes)
    total_proy = proyectar_eom_runrate(total_real, dias_transc, dias_total_mes)

    driver = principal_driver_gap(pd.concat([d_rep, d_srv], ignore_index=True))
    driver_txt = f"Principal desvío: **{driver['KPI']} / {driver['Cat']}** (Gap {money_str(driver['Gap'])})" if driver else "Principal desvío: —"

    st.info(
        f"**Resumen Ejecutivo:** "
        f"Repuestos {pct_str(rep_c)} | "
        f"Servicios {pct_str(srv_c)} | "
        f"Total Postventa {pct_str(total_c)} | "
        f"{driver_txt}"
    )

    rep_month = df_month[(df_month["Tipo_KPI"] == "$") & (df_month["KPI"].str.upper() == "REPUESTOS")].copy()
    rep_month = rep_month[rep_month["Categoria_KPI"].isin(rep_sel)].copy()
    srv_month = df_month[(df_month["Tipo_KPI"] == "$") & (df_month["KPI"].str.upper() == "SERVICIOS")].copy()
    srv_month = srv_month[srv_month["Categoria_KPI"].isin(srv_sel)].copy()
    total_month = pd.concat([rep_month, srv_month], ignore_index=True)

    def macro_block(titulo, real, obj, proy, df_scope_month_for_spark, chart_key):
        c = safe_ratio(real, obj)
        sub = (
            f"Real {money_str(real)} | Obj {money_str(obj)} | "
            f"Proy período (run-rate): {money_str(proy)} | "
            f"Días: {('—' if pd.isna(dias_transc) else int(dias_transc))}/{('—' if pd.isna(dias_total_mes) else int(dias_total_mes))}"
        )
        st.markdown(card_html_base(titulo, money_str(real), sub), unsafe_allow_html=True)
        st.markdown(f"**Cumpl. Acum.:** {pct_str(c)}")
        spark_evolucion(df_scope_month_for_spark, chart_key=chart_key)

    c1, c_mid, c2 = st.columns([1.0, 1.05, 1.0])

    with c1:
        st.markdown("### 🧩 REPUESTOS (P&L)")
        macro_block(
            "Repuestos — Real (Acum.)",
            rep_real,
            rep_obj,
            rep_proy,
            rep_month,
            chart_key="spark_rep_pyl"
        )

    with c_mid:
        st.markdown("### 🧩 TOTAL POSTVENTA (P&L)")
        macro_block(
            "Total Postventa — Real (Acum.)",
            total_real,
            total_obj,
            total_proy,
            total_month,
            chart_key="spark_total_pyl"
        )

    with c2:
        st.markdown("### 🧩 SERVICIOS (P&L)")
        macro_block(
            "Servicios — Real (Acum.)",
            srv_real,
            srv_obj,
            srv_proy,
            srv_month,
            chart_key="spark_srv_pyl"
        )

    st.markdown("---")
    st.markdown("### Cumplimiento por Sucursal — (acumulado)")

    a, b = st.columns(2)
    with a:
        st.markdown("**Repuestos — por sucursal**")
        g = micro_sucursal(apply_obj0_filter(d_rep, show_obj0))
        if g.empty:
            st.info("Sin datos por sucursal.")
        else:
            fig = px.bar(g, x="Cumpl_plot", y="Sucursal", orientation="h", text="label")
            fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
            fig.update_traces(textposition="inside")
            st.plotly_chart(fig, use_container_width=True, key="pyl_rep_sucursal")

    with b:
        st.markdown("**Servicios — por sucursal**")
        g = micro_sucursal(apply_obj0_filter(d_srv, show_obj0))
        if g.empty:
            st.info("Sin datos por sucursal.")
        else:
            fig = px.bar(g, x="Cumpl_plot", y="Sucursal", orientation="h", text="label")
            fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
            fig.update_traces(textposition="inside")
            st.plotly_chart(fig, use_container_width=True, key="pyl_srv_sucursal")

    st.markdown("---")
    st.markdown("### Aperturas — micro (cumplimiento acumulado)")

    l, r = st.columns(2)
    with l:
        st.markdown("**Repuestos — por apertura**")
        g = micro_aperturas(apply_obj0_filter(d_rep, show_obj0))
        if g.empty:
            st.info("Sin datos (revisar aperturas).")
        else:
            fig = px.bar(g, x="Cumpl_plot", y="Categoria_KPI", orientation="h", text="label")
            fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
            fig.update_traces(textposition="inside")
            st.plotly_chart(fig, use_container_width=True, key="pyl_rep_aperturas")

    with r:
        st.markdown("**Servicios — por apertura**")
        g = micro_aperturas(apply_obj0_filter(d_srv, show_obj0))
        if g.empty:
            st.info("Sin datos (revisar aperturas).")
        else:
            fig = px.bar(g, x="Cumpl_plot", y="Categoria_KPI", orientation="h", text="label")
            fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
            fig.update_traces(textposition="inside")
            st.plotly_chart(fig, use_container_width=True, key="pyl_srv_aperturas")

    st.markdown("---")
    st.markdown("## 🎯 Micro — ranking sucursal + apertura")

    cA, cB, cC, cD = st.columns([1.1, 1.2, 1.2, 1.5])
    with cA:
        top_n = st.selectbox("Top N", [5, 10, 15, 20, 30], index=1)
        st.session_state["top_n_rank"] = int(top_n)
    with cB:
        rep_micro_choice = st.selectbox("Repuestos (micro)", ["Todas las aperturas"] + rep_open, index=0)
    with cC:
        srv_micro_choice = st.selectbox("Servicios (micro)", ["Todas las aperturas"] + srv_open, index=0)
    with cD:
        show_zero_rank = st.checkbox("Mostrar 0% (Obj=0 y real=0)", value=False)
        st.session_state["show_zero_rank"] = bool(show_zero_rank)

    rep_rank_base = d_rep.copy()
    if rep_micro_choice != "Todas las aperturas":
        rep_rank_base = rep_rank_base[rep_rank_base["Categoria_KPI"] == rep_micro_choice].copy()

    srv_rank_base = d_srv.copy()
    if srv_micro_choice != "Todas las aperturas":
        srv_rank_base = srv_rank_base[srv_rank_base["Categoria_KPI"] == srv_micro_choice].copy()

    rr, ss = st.columns(2)
    with rr:
        st.markdown("### Repuestos — sucursal + apertura (micro)")
        g = ranking_sucursal_apertura_micro(rep_rank_base, top_n=top_n, show_zero=show_zero_rank)
        if g.empty:
            st.info("Sin ranking.")
        else:
            fig = px.bar(g, x="Cumpl_plot", y="key", orientation="h", text="label")
            fig.update_layout(height=460, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
            fig.update_traces(textposition="inside")
            st.plotly_chart(fig, use_container_width=True, key="pyl_rep_micro")

    with ss:
        st.markdown("### Servicios — sucursal + apertura (micro)")
        g = ranking_sucursal_apertura_micro(srv_rank_base, top_n=top_n, show_zero=show_zero_rank)
        if g.empty:
            st.info("Sin ranking.")
        else:
            fig = px.bar(g, x="Cumpl_plot", y="key", orientation="h", text="label")
            fig.update_layout(height=460, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
            fig.update_traces(textposition="inside")
            st.plotly_chart(fig, use_container_width=True, key="pyl_srv_micro")

# ============================================================
# TAB 2 — KPIs resto
# ============================================================
with tab2:
    st.markdown("## 📌 KPIs (resto) — Macro → Micro")
    st.markdown("---")

    resto = df_cut[~df_cut["KPI"].str.upper().isin(["REPUESTOS", "SERVICIOS"])].copy()
    resto = apply_obj0_filter(resto, show_obj0)

    kpis_resto = sorted(resto["KPI"].unique().tolist())
    if not kpis_resto:
        st.info("No hay KPIs (resto) con Obj>0 en este corte.")
    else:
        kpi_sel = st.selectbox("Elegí un KPI (resto)", kpis_resto)

        x = resto[resto["KPI"] == kpi_sel].copy()
        tipos = sorted(x["Tipo_KPI"].unique().tolist())

        for t in tipos:
            xt = x[x["Tipo_KPI"] == t].copy()
            real = xt["Real_val"].sum()
            obj = xt["Obj_val"].sum()
            c = safe_ratio(real, obj)

            st.markdown(
                card_html_base(
                    f"{kpi_sel} ({t}) — Cumplimiento (Acum.)",
                    (money_str(real) if t == "$" else qty_str(real)),
                    f"Real {(money_str(real) if t == '$' else qty_str(real))} | Obj {(money_str(obj) if t == '$' else qty_str(obj))}"
                ),
                unsafe_allow_html=True
            )
            st.markdown(f"**Cumpl. Acum.:** {pct_str(c)}")

            g = xt.groupby("Sucursal", as_index=False).agg(
                Real=("Real_val", "sum"),
                Obj=("Obj_val", "sum")
            )
            g["Cumpl"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
            g = g[~g["Cumpl"].isna()].copy().sort_values("Cumpl", ascending=False)
            g = apply_cap_visual(g, cap_on, cap_val)

            g["label"] = g.apply(
                lambda r: f"{pct_str(r['Cumpl'])} | {(money_str(r['Real']) if t == '$' else qty_str(r['Real']))}/{(money_str(r['Obj']) if t == '$' else qty_str(r['Obj']))}",
                axis=1
            )

            st.markdown("### Ranking por sucursal — este KPI")
            if g.empty:
                st.info("Sin ranking (Obj=0 o sin datos).")
            else:
                fig = px.bar(g, x="Cumpl_plot", y="Sucursal", orientation="h", text="label")
                fig.update_layout(height=420, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Cumplimiento (visual)")
                fig.update_traces(textposition="inside")
                st.plotly_chart(fig, use_container_width=True, key=f"kpi_resto_{kpi_sel}_{t}")

        st.markdown("---")
        with st.expander("🔎 Auditoría (KPIs resto)", expanded=False):
            detail = x.copy().sort_values(["Mes", "Semana_Num", "Sucursal", "KPI", "Categoria_KPI"])
            st.dataframe(detail, use_container_width=True, hide_index=True)

# ============================================================
# TAB 3 — Gestión
# ============================================================
with tab3:
    st.markdown("## 🧪 Gestión (desvíos)")
    st.markdown("---")

    suc_g_opts = ["TODAS (Consolidado)"] + sorted(df["Sucursal"].dropna().unique().tolist())
    suc_g = st.selectbox("Sucursal (Gestión)", suc_g_opts, index=0)

    d = df[df["Mes"].isin(meses_sel)].copy()
    if suc_g != "TODAS (Consolidado)":
        d = d[d["Sucursal"] == suc_g].copy()
    elif not selected_all_suc:
        d = d[d["Sucursal"].isin(selected_sucursales_real)].copy()

    d = d[d["Semana_Num"].isin(semanas_sel)].copy()
    d = apply_obj0_filter(d, show_obj0)

    g = d.groupby(["KPI", "Categoria_KPI", "Tipo_KPI"], as_index=False).agg(
        Real=("Real_val", "sum"),
        Obj=("Obj_val", "sum")
    )
    g["Gap"] = g["Obj"] - g["Real"]
    g["Cumpl"] = g.apply(lambda r: safe_ratio(r["Real"], r["Obj"]), axis=1)
    g = g.sort_values("Gap", ascending=False)

    st.markdown("### Top desvíos (Gap) — Obj - Real")
    if g.empty:
        st.info("Sin desvíos (o todo Obj=0).")
    else:
        show_n = st.selectbox("Top N desvíos", [10, 20, 30, 50], index=1)
        gg = g.head(show_n).copy()
        gg["key"] = gg["KPI"].astype(str) + " — " + gg["Categoria_KPI"].astype(str) + " (" + gg["Tipo_KPI"].astype(str) + ")"
        gg["label"] = gg.apply(
            lambda r: f"Gap {(money_str(r['Gap']) if r['Tipo_KPI'] == '$' else qty_str(r['Gap']))} | {pct_str(r['Cumpl'])}",
            axis=1
        )

        fig = px.bar(gg, x="Gap", y="key", orientation="h", text="label")
        fig.update_layout(height=520, margin=dict(l=10, r=10, t=10, b=10), xaxis_title="Gap (Obj - Real)")
        fig.update_traces(textposition="inside")
        st.plotly_chart(fig, use_container_width=True, key="gestion_desvios_gap")

        with st.expander("🔎 Auditoría (Gestión)", expanded=False):
            st.dataframe(g, use_container_width=True, hide_index=True)

# ============================================================
# TAB 4 — Hitos del mes
# ============================================================
with tab4:
    st.markdown("## 🗓️ Hitos del mes — Resumen del mes")
    st.caption(f"Mes(es) seleccionados: **{labels_from_mes_keys(meses_sel)}**")
    st.markdown("---")

    if RESUMEN_SHEET_FOUND is None:
        st.warning("No pude encontrar la hoja 'Resumen del mes' por nombre.")
        with st.expander("Ver pestañas detectadas", expanded=True):
            st.write(SHEET_NAMES)
        st.stop()

    st.success(f"✅ Hoja encontrada: **{RESUMEN_SHEET_FOUND}**")

    hitos_mes = filter_hitos_by_month(df_resumen_mes, meses_sel)

    if hitos_mes is None or hitos_mes.empty:
        st.info("La hoja existe, pero no tiene contenido para los meses seleccionados.")
    else:
        st.dataframe(hitos_mes, use_container_width=True, hide_index=True)

# ============================================================
# TAB 5 — ÓRDENES ABIERTAS
# ============================================================
with tab5:
    abiertas_std = build_operational_standard(df_abiertas_raw, "Abiertas")
    render_operational_tab(
        abiertas_std,
        "🔧 Órdenes Abiertas",
        "Vehículos actualmente en taller con trabajos o reparaciones pendientes. Foco de gestión: destrabar backlog, acelerar terminación de trabajos y cierre de órdenes.",
        key_prefix="abiertas",
        enable_asesor_filter=True
    )

# ============================================================
# TAB 6 — PENDIENTES FACTURACIÓN
# ============================================================
with tab6:
    pfact_std = build_operational_standard(df_pfact_raw, "Pendientes Fact")
    render_operational_tab(
        pfact_std,
        "🧾 Pendientes de Facturación",
        "Órdenes finalizadas que todavía no se transformaron en facturación/cobro. Foco de gestión: conversión a caja y cierre administrativo.",
        key_prefix="pfact",
        enable_asesor_filter=True
    )

# ============================================================
# TAB 7 — PRESUPUESTOS
# ============================================================
with tab7:
    presup_std = build_operational_standard(df_presup_raw, "Presupuestos")

    if presup_std is None or presup_std.empty:
        st.markdown("## 💬 Presupuestos Pendientes")
        st.info("No hay registros en esta hoja o no se pudieron interpretar columnas con contenido útil.")
    else:
        estados_disp = sorted([e for e in presup_std["Estado"].dropna().astype(str).str.strip().unique().tolist() if e != ""])
        estado_sel = st.multiselect(
            "Filtrar estado (columna H)",
            estados_disp,
            default=estados_disp,
            key="presup_estado"
        )
        presup_f = presup_std[presup_std["Estado"].astype(str).isin(estado_sel)].copy() if estados_disp else presup_std.copy()

        render_operational_tab(
            presup_f,
            "💬 Presupuestos Pendientes",
            "Presupuestos aún no aprobados por el cliente. Foco de gestión: seguimiento comercial, recuperación de ventas y priorización por importe/antigüedad.",
            key_prefix="presup",
            enable_asesor_filter=False
        )

# ============================================================
# TAB 8 — CIERRE GAP
# ============================================================
with tab8:
    render_tab_cierre_gap()
